#!/usr/bin/env python3
"""
Stage 6 — Screen paper abstracts against the research prompt using Claude Sonnet.

Rates each paper's relevance (1-10), classifies as theoretical/empirical,
identifies methodology, and tags relationship to the user's research.

Usage:
    python lit_screen.py --input stage5_merged.json --query "research prompt" -o stage6_screened.json

Requires: pip install aiohttp openpyxl
API keys: ANTHROPIC_API_KEY
"""

import argparse
import asyncio
import json
import logging
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# Fix Windows console encoding
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# Load ~/.lit-review-pipeline.env if present (portable key store)
_env_file = Path.home() / ".lit-review-pipeline.env"
if _env_file.exists():
    try:
        from dotenv import load_dotenv
        load_dotenv(_env_file, override=False)
    except ImportError:
        pass

try:
    import aiohttp
except ImportError:
    print("Error: aiohttp required. Install with: pip install aiohttp")
    sys.exit(1)


# ── Logging & Heartbeat ─────────────────────────────────────────────────────

def _setup_logging(log_path: Path) -> None:
    """Configure file + console logging."""
    logger = logging.getLogger("lit-screen")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fh = logging.FileHandler(str(log_path), mode="a", encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s",
                                       datefmt="%H:%M:%S"))
    logger.addHandler(fh)

    ch = logging.StreamHandler(sys.stderr)
    ch.setLevel(logging.WARNING)
    ch.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
    logger.addHandler(ch)


# Global heartbeat state — updated from async screening loop
_heartbeat_state: dict = {
    "path": None,
    "start_time": None,
}


def _write_heartbeat(completed: int, total: int, status: str, avg_score: float = 0.0) -> None:
    """Write heartbeat JSON (direct overwrite, Dropbox-safe)."""
    hb_path = _heartbeat_state.get("path")
    if not hb_path:
        return
    hb = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "start_time": _heartbeat_state.get("start_time", ""),
        "status": status,
        "completed": completed,
        "total": total,
        "avg_score": round(avg_score, 2),
    }
    try:
        hb_path.write_text(json.dumps(hb, indent=2), encoding="utf-8")
    except PermissionError:
        pass  # Dropbox lock — skip this heartbeat


# ── Constants ────────────────────────────────────────────────────────────────

ANTHROPIC_API_URL = "https://api.anthropic.com/v1/messages"

SCREENING_SYSTEM_PROMPT = """\
You are screening academic papers for a literature review. Given a research \
prompt and a paper's metadata, rate the paper's relevance and classify it.

Respond in EXACTLY this JSON format (no markdown, no extra text):
{
  "relevance_score": <int 1-10>,
  "rationale": "<one sentence explaining the relevance rating>",
  "paper_type": "<theoretical|empirical>",
  "identification_strategy": "<natural experiment|IV|DiD|RDD|structural|descriptive|N/A>",
  "relationship": "<foundational/must-cite|same method different context|same context different method|direct competitor|methodological reference|tangential>"
}"""


# ── Per-paper screening ─────────────────────────────────────────────────────

def _build_user_message(paper: dict, query: str) -> str:
    """Build the user message for screening a single paper."""
    title = paper.get("title", "N/A")
    authors = paper.get("authors", "N/A")
    year = paper.get("year", "N/A")
    journal = paper.get("journal", "N/A")
    abstract = (paper.get("abstract") or "").strip()
    if abstract:
        abstract_line = f"Abstract: {abstract}"
    else:
        abstract_line = ("Abstract: (none available — assess relevance from the title, "
                         "venue, year, and authors, and score conservatively, typically no "
                         "higher than 6 without an abstract.)")

    return (
        f"Research prompt: {query}\n\n"
        f"Paper:\n"
        f"Title: {title}\n"
        f"Authors: {authors}\n"
        f"Year: {year}\n"
        f"Journal: {journal}\n"
        f"{abstract_line}"
    )


def _parse_screening_response(raw: str) -> dict:
    """Parse the structured JSON response from Claude."""
    # Strip markdown fences if present
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)

    parsed = json.loads(text)
    return {
        "screening_score": int(parsed.get("relevance_score", 0)),
        "screening_rationale": str(parsed.get("rationale", "")),
        "paper_type": str(parsed.get("paper_type", "N/A")),
        "identification_strategy": str(parsed.get("identification_strategy", "N/A")),
        "relationship": str(parsed.get("relationship", "N/A")),
    }


def _no_abstract_result() -> dict:
    return {
        "screening_score": 0,
        "screening_rationale": "No abstract available",
        "paper_type": "N/A",
        "identification_strategy": "N/A",
        "relationship": "N/A",
    }


def _error_result(error: str) -> dict:
    return {
        "screening_score": 0,
        "screening_rationale": f"Screening failed: {error}",
        "paper_type": "N/A",
        "identification_strategy": "N/A",
        "relationship": "N/A",
    }


def _result_to_fields(r: dict) -> dict:
    """Map an Opus screening result (raw or canonical keys) to the merged schema."""
    score = r.get("relevance_score", r.get("screening_score", 0))
    try:
        score = int(score)
    except (TypeError, ValueError):
        score = 0
    return {
        "screening_score": score,
        "screening_rationale": str(r.get("rationale", r.get("screening_rationale", ""))),
        "paper_type": str(r.get("paper_type", "N/A")),
        "identification_strategy": str(r.get("identification_strategy", "N/A")),
        "relationship": str(r.get("relationship", "N/A")),
    }


async def _screen_one(
    session: aiohttp.ClientSession,
    sem: asyncio.Semaphore,
    paper: dict,
    query: str,
    api_key: str,
    model: str,
) -> dict:
    """Screen a single paper against the research query.

    Papers with no abstract are still screened on title/venue/authors (with a
    conservative-scoring instruction in the user message) rather than auto-dropped
    at score 0, so a clearly on-topic working paper or citation-chain hit can
    still pass the relevance filter.
    """
    user_msg = _build_user_message(paper, query)

    body = {
        "model": model,
        "max_tokens": 300,
        "temperature": 0,
        "system": SCREENING_SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": user_msg}],
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "Content-Type": "application/json",
    }

    max_retries = 5
    for attempt in range(max_retries):
        async with sem:
            try:
                async with session.post(
                    ANTHROPIC_API_URL,
                    json=body,
                    headers=headers,
                    timeout=aiohttp.ClientTimeout(total=60),
                ) as resp:
                    if resp.status == 429 or resp.status == 529:
                        retry_after = resp.headers.get("retry-after")
                        wait = float(retry_after) if retry_after else (2 ** attempt + 1)
                        if attempt < max_retries - 1:
                            await asyncio.sleep(wait)
                            continue
                        err_text = await resp.text()
                        return _error_result(f"HTTP {resp.status}: {err_text[:100]}")
                    if resp.status != 200:
                        err_text = await resp.text()
                        return _error_result(f"HTTP {resp.status}: {err_text[:100]}")
                    data = await resp.json()
                    raw = data["content"][0]["text"].strip()
                    return _parse_screening_response(raw)
            except json.JSONDecodeError as e:
                return _error_result(f"JSON parse error: {e}")
            except Exception as e:
                if attempt < max_retries - 1:
                    await asyncio.sleep(2 ** attempt)
                    continue
                return _error_result(str(e)[:100])
    return _error_result("Max retries exceeded")


# ── Batch screening ─────────────────────────────────────────────────────────

async def _screen_batch(
    papers: list[dict],
    query: str,
    api_key: str,
    model: str,
    concurrency: int,
    checkpoint_path: Path | None = None,
    checkpoint_every: int = 25,
) -> list[dict]:
    """Screen all papers concurrently with a semaphore and periodic checkpointing.

    If checkpoint_path is given, saves partial results every checkpoint_every papers
    so that progress survives crashes. The checkpoint is a JSON file with the same
    structure as the final output (papers list with screening fields merged in).
    """
    sem = asyncio.Semaphore(concurrency)
    connector = aiohttp.TCPConnector(limit=concurrency)
    async with aiohttp.ClientSession(connector=connector) as session:
        tasks = [
            _screen_one(session, sem, p, query, api_key, model)
            for p in papers
        ]

        # Gather with progress reporting + checkpointing
        total = len(tasks)
        results: list[dict] = [{}] * total
        completed = 0
        score_sum = 0.0
        _last_checkpoint = 0

        _logger = logging.getLogger("lit-screen")

        async def _run_with_progress(idx: int, coro):
            nonlocal completed, score_sum, _last_checkpoint
            result = await coro
            results[idx] = result
            completed += 1
            score_sum += result.get("screening_score", 0)
            if completed % 10 == 0 or completed == total:
                avg = score_sum / completed
                msg = f"[{completed}/{total}] scored — avg relevance: {avg:.1f}"
                print(f"  {msg}")
                _logger.info(msg)
                _write_heartbeat(completed, total, "running", avg)

            # Periodic checkpoint
            if checkpoint_path and (completed - _last_checkpoint) >= checkpoint_every:
                _last_checkpoint = completed
                _save_checkpoint(papers, results, checkpoint_path)
                print(f"  [checkpoint] saved {completed}/{total} to {checkpoint_path.name}")
                _logger.info(f"Checkpoint saved: {completed}/{total}")

            return idx, result

        progress_tasks = [
            _run_with_progress(i, t) for i, t in enumerate(tasks)
        ]
        for idx, result in await asyncio.gather(*progress_tasks):
            results[idx] = result

    # Final checkpoint
    if checkpoint_path:
        _save_checkpoint(papers, results, checkpoint_path)

    return results


def _safe_write_text(path: Path, text: str) -> None:
    """Write text to path, retrying once after a short wait on PermissionError.

    Uses direct overwrite (not atomic rename) because Dropbox/OneDrive sync can
    hold locks that cause PermissionError; the retry-after-sleep mirrors
    _save_checkpoint so final artifacts survive a transient sync lock.
    """
    try:
        path.write_text(text, encoding="utf-8")
    except PermissionError:
        # Dropbox/sync lock — retry once after short wait
        import time
        time.sleep(0.5)
        try:
            path.write_text(text, encoding="utf-8")
        except PermissionError:
            logging.getLogger("lit-screen").warning(f"Write failed (Dropbox lock): {path}")


def _save_checkpoint(papers: list[dict], results: list[dict], path: Path) -> None:
    """Save partial screening results as a checkpoint file.

    Uses direct overwrite instead of atomic rename because Dropbox/OneDrive
    sync can hold locks that cause PermissionError on os.replace().
    """
    merged = []
    for paper, result in zip(papers, results):
        p = dict(paper)
        if result:
            p.update(result)
        merged.append(p)
    _safe_write_text(path, json.dumps(merged, indent=2, ensure_ascii=False))


def _load_checkpoint(papers: list[dict], checkpoint_path: Path) -> tuple[list[dict], list[int]]:
    """Load a checkpoint and identify which papers still need screening.

    Returns (papers_with_partial_results, indices_to_screen).
    Papers that were successfully screened (have a valid screening_score > 0
    and no error in rationale) are kept; others are marked for re-screening.
    """
    checkpoint_data = json.loads(checkpoint_path.read_text(encoding="utf-8"))
    if not isinstance(checkpoint_data, list) or len(checkpoint_data) != len(papers):
        print(f"  Warning: checkpoint has {len(checkpoint_data)} papers, input has {len(papers)} — ignoring checkpoint")
        return papers, list(range(len(papers)))

    to_screen = []
    for i, cp in enumerate(checkpoint_data):
        score = cp.get("screening_score", 0)
        rationale = cp.get("screening_rationale", "")
        if score > 0 and not rationale.startswith("Screening failed:"):
            # Already successfully screened — merge result into paper
            papers[i].update({
                "screening_score": cp["screening_score"],
                "screening_rationale": cp.get("screening_rationale", ""),
                "paper_type": cp.get("paper_type", "N/A"),
                "identification_strategy": cp.get("identification_strategy", "N/A"),
                "relationship": cp.get("relationship", "N/A"),
            })
        else:
            to_screen.append(i)

    return papers, to_screen


# ── RIS Writer ───────────────────────────────────────────────────────────────

def write_ris(papers: list[dict], path: Path) -> None:
    """Write papers in RIS format."""
    lines = []
    for p in papers:
        lines.append("TY  - JOUR")
        if p.get("title"):
            lines.append(f"T1  - {p['title']}")
        for author in (p.get("authors") or "").split(", "):
            author = author.strip()
            if author:
                lines.append(f"AU  - {author}")
        if p.get("journal"):
            lines.append(f"JO  - {p['journal']}")
        if p.get("year"):
            lines.append(f"PY  - {p['year']}")
        if p.get("doi"):
            doi = re.sub(r"^https?://doi\.org/", "", p["doi"])
            lines.append(f"DO  - {doi}")
        if p.get("abstract"):
            lines.append(f"AB  - {p['abstract']}")
        if p.get("url"):
            lines.append(f"UR  - {p['url']}")
        lines.append("ER  -")
        lines.append("")

    _safe_write_text(path, "\n".join(lines))


# ── BibTeX Writer ────────────────────────────────────────────────────────────

def _make_cite_key(paper: dict) -> str:
    """Generate author_year citation key, e.g. 'Smith2024'."""
    authors = (paper.get("authors") or "").strip()
    year = str(paper.get("year") or "").strip()

    # Extract first author surname
    if authors:
        first = authors.split(",")[0].strip()
        # Handle "Last, First" or "First Last" formats
        parts = first.split()
        surname = parts[-1] if parts else "Unknown"
    else:
        surname = "Unknown"

    # Clean surname: keep only alphanumeric
    surname = re.sub(r"[^a-zA-Z]", "", surname)
    if not surname:
        surname = "Unknown"

    return f"{surname}{year}"


def write_bib(papers: list[dict], path: Path, min_score: int = 4) -> int:
    """Write papers with screening_score >= min_score in BibTeX format.

    Returns the number of entries written.
    """
    eligible = [p for p in papers if p.get("screening_score", 0) >= min_score]
    eligible.sort(key=lambda p: p.get("screening_score", 0), reverse=True)

    # Deduplicate citation keys by appending a/b/c suffixes
    key_counts: dict[str, int] = {}
    entries = []

    for p in eligible:
        base_key = _make_cite_key(p)
        count = key_counts.get(base_key, 0)
        key_counts[base_key] = count + 1
        cite_key = base_key if count == 0 else f"{base_key}{chr(97 + count)}"

        lines = [f"@article{{{cite_key},"]

        title = p.get("title", "")
        if title:
            lines.append(f"  title = {{{title}}},")

        authors = p.get("authors", "")
        if authors:
            # BibTeX wants "Last, First and Last, First" format
            # Input is typically "First Last, First Last, ..."
            author_list = [a.strip() for a in authors.split(",") if a.strip()]
            lines.append(f"  author = {{{' and '.join(author_list)}}},")

        year = p.get("year", "")
        if year:
            lines.append(f"  year = {{{year}}},")

        journal = p.get("journal", "")
        if journal:
            lines.append(f"  journal = {{{journal}}},")

        doi = (p.get("doi") or "").replace("https://doi.org/", "")
        if doi:
            lines.append(f"  doi = {{{doi}}},")

        abstract = p.get("abstract", "")
        if abstract:
            # Escape special BibTeX characters in abstract
            safe = abstract.replace("&", r"\&").replace("%", r"\%")
            lines.append(f"  abstract = {{{safe}}},")

        url = p.get("url", "")
        if url:
            lines.append(f"  url = {{{url}}},")

        # Add screening metadata as custom fields
        lines.append(f"  screening_score = {{{p.get('screening_score', '')}}},")
        lines.append(f"  paper_type = {{{p.get('paper_type', '')}}},")
        lines.append(f"  relationship = {{{p.get('relationship', '')}}},")

        lines.append("}")
        entries.append("\n".join(lines))

    _safe_write_text(path, "\n\n".join(entries) + "\n")
    return len(entries)


# ── Abstract type classification ────────────────────────────────────────────

def _classify_abstract_type(paper: dict) -> str:
    """Classify whether a paper's abstract is a real abstract or a snippet.

    Returns: 'abstract', 'snippet', or 'none'.
    """
    abstract = (paper.get("abstract") or "").strip()
    if not abstract:
        return "none"

    # Explicitly tagged as snippet by the backfill cascade
    if paper.get("_abstract_source") == "google_scholar_snippet":
        return "snippet"

    # Supplementary-search papers use Google Scholar snippets as "abstracts"
    # These are typically ~150-200 chars with ellipsis markers
    source = paper.get("source", "") or ""
    if source in ("ssrn", "nber", "heinonline", "forthcoming"):
        # Heuristic: real abstracts are >250 chars; snippets have "…" and are short
        if len(abstract) < 250 and ("…" in abstract or "..." in abstract):
            return "snippet"

    return "abstract"


# ── XLSX Writer ──────────────────────────────────────────────────────────────

def write_xlsx(papers: list[dict], path: Path) -> None:
    """Write screening results to XLSX, sorted by screening_score desc."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    COLS = [
        "title", "authors", "year", "journal", "DOI", "source",
        "screening_score", "paper_type", "identification_strategy",
        "relationship", "screening_rationale", "abstract_type", "abstract",
    ]
    COL_WIDTHS = {
        "title": 60, "authors": 35, "year": 8, "journal": 30,
        "DOI": 28, "source": 16, "screening_score": 14,
        "paper_type": 14, "identification_strategy": 22,
        "relationship": 28, "screening_rationale": 60,
        "abstract_type": 12, "abstract": 80,
    }

    # Sort by screening_score descending
    sorted_papers = sorted(
        papers, key=lambda p: p.get("screening_score", 0), reverse=True
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Screened Papers"

    # Header row
    bold = Font(bold=True)
    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = bold
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 16)

    # Data rows
    wrap = Alignment(wrap_text=True, vertical="top")
    for ri, p in enumerate(sorted_papers, 2):
        for ci, col in enumerate(COLS, 1):
            if col == "DOI":
                val = (p.get("doi") or "").replace("https://doi.org/", "")
            elif col == "source":
                val = p.get("source", "")
                if not val:
                    sources = p.get("sources", [])
                    val = sources[0] if sources else ""
            elif col == "abstract_type":
                val = _classify_abstract_type(p)
            else:
                val = p.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            if col in ("title", "abstract", "screening_rationale"):
                cell.alignment = wrap

    # Auto-filter and freeze panes
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(sorted_papers) + 1}"
    ws.freeze_panes = "A2"

    try:
        wb.save(path)
    except PermissionError:
        # Dropbox/sync lock — retry once after short wait
        import time
        time.sleep(0.5)
        wb.save(path)


# ── Summary ──────────────────────────────────────────────────────────────────

def print_summary(papers: list[dict]) -> None:
    """Print screening summary statistics."""
    total = len(papers)
    scores = [p.get("screening_score", 0) for p in papers]
    avg = sum(scores) / total if total else 0

    ge7 = sum(1 for s in scores if s >= 7)
    ge5 = sum(1 for s in scores if s >= 5)

    # Type counts
    type_counts: dict[str, int] = {}
    for p in papers:
        pt = p.get("paper_type", "N/A")
        type_counts[pt] = type_counts.get(pt, 0) + 1

    # Relationship counts
    rel_counts: dict[str, int] = {}
    for p in papers:
        rel = p.get("relationship", "N/A")
        rel_counts[rel] = rel_counts.get(rel, 0) + 1

    print(f"\n=== Screening Summary ===")
    print(f"  Total papers:     {total}")
    print(f"  Avg relevance:    {avg:.1f}")
    print(f"  Score >= 7:       {ge7:>4d} papers ({100*ge7/total:.1f}%)" if total else "")
    print(f"  Score >= 5:       {ge5:>4d} papers ({100*ge5/total:.1f}%)" if total else "")
    print()
    type_str = ", ".join(f"{k}={v}" for k, v in sorted(type_counts.items()))
    print(f"  By type: {type_str}")
    rel_str = ", ".join(f"{k}={v}" for k, v in sorted(rel_counts.items()))
    print(f"  By relationship: {rel_str}")


def _write_screen_outputs(papers: list[dict], output_path: Path) -> None:
    """Write the six Stage-6 artifacts (JSON x2, XLSX x2, RIS, BibTeX).

    Shared by the autonomous API path and the agent-driven --ingest-results path.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    _safe_write_text(output_path, json.dumps(papers, indent=2, ensure_ascii=False))
    print(f"\nSaved {len(papers)} papers to {output_path}")

    filtered = [p for p in papers if p.get("screening_score", 0) >= 4]
    # Derive a distinct "filtered" stem. If the output stem has no "screened" to
    # rewrite (e.g. -o screen_out.json), append a suffix so the filtered files never
    # overwrite the full screened JSON/XLSX.
    filtered_stem = output_path.stem.replace("screened", "filtered")
    if filtered_stem == output_path.stem:
        filtered_stem = f"{output_path.stem}_filtered"
    filtered_path = output_path.parent / f"{filtered_stem}.json"
    _safe_write_text(filtered_path, json.dumps(filtered, indent=2, ensure_ascii=False))
    print(f"Saved {len(filtered)} filtered papers (score >= 4) to {filtered_path}")

    filtered_xlsx_path = output_path.parent / f"{filtered_stem}.xlsx"
    try:
        write_xlsx(filtered, filtered_xlsx_path)
        print(f"Filtered XLSX saved to {filtered_xlsx_path}")
    except ImportError:
        print("Warning: openpyxl not installed, skipping filtered XLSX export")
    except Exception as e:
        print(f"Warning: filtered XLSX export failed: {e}")

    xlsx_path = output_path.parent / f"{output_path.stem}.xlsx"
    try:
        write_xlsx(papers, xlsx_path)
        print(f"XLSX saved to {xlsx_path}")
    except ImportError:
        print("Warning: openpyxl not installed, skipping XLSX export")
    except Exception as e:
        print(f"Warning: XLSX export failed: {e}")

    ris_path = output_path.parent / f"{output_path.stem}.ris"
    write_ris(papers, ris_path)
    print(f"RIS saved to {ris_path}")

    bib_path = output_path.parent / f"{output_path.stem}.bib"
    bib_count = write_bib(papers, bib_path, min_score=4)
    print(f"BibTeX saved to {bib_path} ({bib_count} entries with score >= 4)")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Stage 6 — Screen paper abstracts against research prompt",
    )
    parser.add_argument("--input", required=True, help="Input JSON from Stage 5")
    parser.add_argument("--query", default="",
                        help="Research query/prompt to screen against (required except with --ingest-results)")
    parser.add_argument("-o", "--output", default="stage6_screened.json",
                        help="Output JSON path (default: stage6_screened.json)")
    parser.add_argument("--model", default="claude-sonnet-4-6",
                        help="Anthropic model ID (default: claude-sonnet-4-6)")
    parser.add_argument("--concurrency", type=int, default=5,
                        help="Max simultaneous API requests (default: 5)")
    parser.add_argument("--resume", action="store_true",
                        help="Resume from checkpoint or existing output: skip already-screened papers")
    parser.add_argument("--checkpoint-every", type=int, default=25,
                        help="Save checkpoint every N papers (default: 25)")

    parser.add_argument("--log-dir", default=None,
                        help="Directory for log + heartbeat files (default: same as output)")
    parser.add_argument("--emit-tasks", metavar="PATH",
                        help="Agent-driven mode: write per-paper screening tasks to PATH and exit "
                             "(no API call). Score them with Opus, then pass --ingest-results.")
    parser.add_argument("--ingest-results", metavar="PATH",
                        help="Agent-driven mode: read Opus screening results from PATH, merge, and "
                             "write all Stage-6 outputs (no API call).")

    args = parser.parse_args()

    # Validate input
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: input file not found: {input_path}")
        sys.exit(1)

    # Load papers (needed by every mode)
    papers = json.loads(input_path.read_text(encoding="utf-8"))
    if not isinstance(papers, list):
        print("Error: input JSON must be an array of paper objects")
        sys.exit(1)

    # ── Agent-driven emit: write per-paper screening tasks for Opus, then stop ──
    if args.emit_tasks:
        if not args.query:
            print("Error: --query is required with --emit-tasks")
            sys.exit(1)
        tasks = [
            {"index": i, "user_message": _build_user_message(p, args.query)}
            for i, p in enumerate(papers)
        ]
        payload = {"system_prompt": SCREENING_SYSTEM_PROMPT, "query": args.query, "tasks": tasks}
        Path(args.emit_tasks).write_text(
            json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"[SCREEN] Emitted {len(tasks)} screening tasks to {args.emit_tasks}")
        return

    # ── Agent-driven ingest: merge Opus screening results and write outputs ──
    if args.ingest_results:
        results = json.loads(Path(args.ingest_results).read_text(encoding="utf-8"))
        applied = 0
        for r in results:
            try:
                idx = int(r["index"])
            except (KeyError, ValueError, TypeError):
                continue
            if 0 <= idx < len(papers):
                papers[idx].update(_result_to_fields(r))
                applied += 1
        for p in papers:
            if "screening_score" not in p:
                p.update(_error_result("no screening result returned"))
        print_summary(papers)
        _write_screen_outputs(papers, Path(args.output))
        print(f"[SCREEN] Ingested {applied} of {len(papers)} results")
        return

    # ── Autonomous path (in-script Sonnet API screening) ──
    if not args.query:
        print("Error: --query is required (except with --ingest-results)")
        sys.exit(1)

    # Setup logging + heartbeat directory
    output_path_early = Path(args.output)
    log_dir = Path(args.log_dir) if args.log_dir else output_path_early.parent
    log_dir.mkdir(parents=True, exist_ok=True)

    log_path = log_dir / "stage6_screen.log"
    _setup_logging(log_path)
    _logger = logging.getLogger("lit-screen")
    _logger.info(f"=== Stage 6 started at {datetime.now().isoformat(timespec='seconds')} ===")
    _logger.info(f"Input: {args.input}, Model: {args.model}, Concurrency: {args.concurrency}")

    # Start heartbeat
    heartbeat_path = log_dir / "stage6_heartbeat.json"
    _heartbeat_state["path"] = heartbeat_path
    _heartbeat_state["start_time"] = datetime.now().isoformat(timespec="seconds")
    _write_heartbeat(0, 0, "starting")

    # Load API key
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        print("Error: ANTHROPIC_API_KEY environment variable not set")
        sys.exit(1)

    # Checkpoint path: same directory as output, with .checkpoint.json suffix
    output_path = Path(args.output)
    checkpoint_path = output_path.parent / f"{output_path.stem}.checkpoint.json"

    # Resume mode: load from checkpoint or existing output
    if args.resume and (checkpoint_path.exists() or output_path.exists()):
        # Prefer checkpoint over finished output (checkpoint is more granular)
        resume_source = checkpoint_path if checkpoint_path.exists() else output_path
        papers, to_screen = _load_checkpoint(papers, resume_source)
        already_done = len(papers) - len(to_screen)

        print(f"=== Lit-Screen (Stage 6) — RESUME ===")
        print(f"Loaded checkpoint: {resume_source.name}")
        print(f"Already screened: {already_done} papers")
        print(f"Remaining: {len(to_screen)} papers to screen")
        if not to_screen:
            print("Nothing to re-screen. All papers already scored.")
        else:
            print(f"Model: {args.model}")
            print(f"Concurrency: {args.concurrency}")
            print()
            print("Screening remaining papers...")

            remaining_papers = [papers[i] for i in to_screen]
            results = asyncio.run(
                _screen_batch(
                    remaining_papers, args.query, api_key, args.model,
                    args.concurrency, checkpoint_path, args.checkpoint_every,
                )
            )
            for idx, result in zip(to_screen, results):
                papers[idx].update(result)
    else:
        print(f"=== Lit-Screen (Stage 6) ===")
        print(f"Input: {input_path} ({len(papers)} papers)")
        print(f"Model: {args.model}")
        print(f"Concurrency: {args.concurrency}")
        print(f"Checkpoint every: {args.checkpoint_every} papers")
        print()
        print("Screening papers against research prompt...")

        # Run screening with checkpointing
        results = asyncio.run(
            _screen_batch(
                papers, args.query, api_key, args.model, args.concurrency,
                checkpoint_path, args.checkpoint_every,
            )
        )

        # Merge results into papers
        for paper, result in zip(papers, results):
            paper.update(result)

    # Print summary
    print_summary(papers)

    # Write outputs (shared writer; also used by --ingest-results)
    _write_screen_outputs(papers, output_path)

    print(f"\nResult: {len(papers)} unique papers screened")

    # Final heartbeat + log
    scores = [p.get("screening_score", 0) for p in papers]
    avg = sum(scores) / len(scores) if scores else 0
    _write_heartbeat(len(papers), len(papers), "complete", avg)
    _logger = logging.getLogger("lit-screen")
    _logger.info(f"=== Stage 6 completed at {datetime.now().isoformat(timespec='seconds')} ===")
    _logger.info(f"Result: {len(papers)} papers screened, avg score {avg:.1f}")

    # Clean up checkpoint after successful completion
    if checkpoint_path.exists():
        checkpoint_path.unlink()
        print(f"Checkpoint removed: {checkpoint_path.name}")


if __name__ == "__main__":
    main()
